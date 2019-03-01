# coding=gbk
from basic_info.Open_DB import MYSQL
from basic_info.get_auth_token import get_headers
from basic_info.setting import MySQL_CONFIG, flow_id_list
from basic_info.format_res import dict_res, get_time
from basic_info.setting import HOST_189
import time, random, requests, xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
from xlutils.copy import copy
import json
import os,threading
import datetime
abs_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))


flow_table = load_workbook(abs_dir("flow_dataset_info.xlsx"))
# info_sheet_names = flow_table.get_sheet_names()
flow_sheet = flow_table.get_sheet_by_name('flow_info')
sheet_rows = flow_sheet.max_row  # 获取行数
flow_id__84_list = []
for row in range(1, sheet_rows):
    if flow_sheet.cell(row=row, column=13).value == 84:
        flow_id__84_list.append(flow_sheet.cell(row=row, column=2).value)
# print(flow_sheet.cell(row=18, column=2).value, flow_sheet.cell(row=18, column=13).value == 84)
print(flow_id__84_list)
print(len(flow_id__84_list))