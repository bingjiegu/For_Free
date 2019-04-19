# coding=gbk
from basic_info.Open_DB import MYSQL
from basic_info.get_auth_token import get_headers
from basic_info.setting import MySQL_CONFIG, flow_id_list
from basic_info.format_res import dict_res, get_time
from basic_info.setting import HOST_189
import time, random, requests, xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
from xlutils.copy import copy
import json
import os,threading
import datetime
abs_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))


class GetCheckoutDataSet(object):
    """该类用来获取批量创建的scheduler对应的execution，执行成功后sink所输出的 dataset id"""

    def __init__(self):
        """初始化数据库连接"""
        self.ms = MYSQL(MySQL_CONFIG["HOST"], MySQL_CONFIG["USER"], MySQL_CONFIG["PASSWORD"], MySQL_CONFIG["DB"])

    def file_flowid_count(self):
        # ---使用openpyxl处理表格 12.26update---
        flow_table = load_workbook(abs_dir("flow_dataset_info.xlsx"))
        # info_sheet_names = flow_table.get_sheet_names()
        info_sheet = flow_table.get_sheet_by_name('flow_info')
        info_sheet_row = info_sheet.max_row  # 获取行数
        flowid_list = []
        for i in range(2, info_sheet_row+1):
            if info_sheet.cell(row=i, column=2).value and len(info_sheet.cell(row=i, column=2).value) > 10:
                flowid_list.append(info_sheet.cell(row=i, column=2).value)
        return flowid_list


    def data_for_create_scheduler(self):
        """获取setting.py中的flow_id_list
        1. 根据flow_id 查找flow_name等信息
        2. 根据查询到的flow信息，拼装创建scheduler所需要使用的data
        """
        print("------开始执行data_for_create_scheduler(self)------\n")
        data_list = []
        # print(info_sheet_row)
        flowid_count = self.file_flowid_count()
        for flow_id in flowid_count:
            # print('------',flow_id, '--------')
            # print(i, flow_id)
            print('flow_id', flow_id)
            try:
                sql = 'select name, flow_type, parameters from merce_flow where id = "%s"' % flow_id
                print(sql)
                flow_info = self.ms.ExecuQuery(sql)
                print('flow_info:', flow_info)
            except Exception as e:
                raise e
            else:
                try:
                    flow_name = flow_info[0]["name"]
                    print('flow_name: ', flow_name)
                    flow_type = flow_info[0]["flow_type"]
                    flow_parameters = flow_info[0]["parameters"]
                    flow_parameters_list = dict_res(flow_parameters)
                    arguments_list = []
                    arguments = {}
                    print('flow_parameters_list:', flow_parameters_list)
                    if flow_parameters_list != [] and flow_parameters_list != None:
                        arguments["name"] = flow_parameters_list[0]["name"]
                        arguments["category"] = flow_parameters_list[0]["category"]
                        arguments["value"] = flow_parameters_list[0]["defaultVal"]
                        arguments["refs"] = flow_parameters_list[0]["refs"]
                        arguments["description"] = flow_parameters_list[0]["description"]
                        arguments_list.append(arguments)
                        print('arguments:', arguments)

                except KeyError as e:
                    raise e
                except IndexError as T:
                    raise T

            data = {
                "configurations": {
                    "arguments": arguments_list,
                    "properties": [
                        {
                            "name": "all.debug",
                            "value": "false"
                        },
                        {
                            "name": "all.dataset-nullable",
                            "value": "false"
                        },
                        {
                            "name": "all.lineage.enable",
                            "value": "true"
                        },
                        {
                            "name": "all.notify-output",
                            "value": "false"
                        },
                        {
                            "name": "all.debug-rows",
                            "value": "20"
                        },
                        {
                            "name": "dataflow.master",
                            "value": "yarn"
                        },
                        {
                            "name": "dataflow.deploy-mode",
                            "value": "client"
                        },
                        {
                            "name": "dataflow.queue",
                            "value": "merce.normal"
                        },
                        {
                            "name": "dataflow.num-executors",
                            "value": "2"
                        },
                        {
                            "name": "dataflow.driver-memory",
                            "value": "512M"
                        },
                        {
                            "name": "dataflow.executor-memory",
                            "value": "1G"
                        },
                        {
                            "name": "dataflow.executor-cores",
                            "value": "2"
                        },
                        {
                            "name": "dataflow.verbose",
                            "value": "true"
                        },
                        {
                            "name": "dataflow.local-dirs",
                            "value": ""
                        },
                        {
                            "name": "dataflow.sink.concat-files",
                            "value": "true"
                        }
                    ],
                    "startTime": get_time()
                },
                "flowId": flow_id,
                "flowName": flow_name,
                "flowType": flow_type,
                "name": flow_name + 'scheduler' + str(random.randint(0, 9999))+str(random.randint(0, 9999)),
                "schedulerId": "once",
                "source": "rhinos"
            }

            data_list.append(data)
        print("------data_for_create_scheduler(self)执行结束------\n")
        return data_list

    def create_new_scheduler(self):
        """
        该方法使用data_for_create_scheduler()返回的data_list批量创建scheduler，
        并返回scheduler_id_list， 供get_execution_info(self)调用
        """
        print("------开始执行create_new_scheduler(self)------\n")
        from basic_info.url_info import create_scheduler_url
        scheduler_id_list = []
        scheduler_number = 1
        for data in self.data_for_create_scheduler():
            res = requests.post(url=create_scheduler_url, headers=get_headers(), json=data)
            print('第%d 个scheduler' % scheduler_number)
            scheduler_number += 1
            time.sleep(2)
            print(res.status_code, res.text)
            if res.status_code == 201 and res.text:
                scheduler_id_format = dict_res(res.text)
                try:
                    scheduler_id = scheduler_id_format["id"]
                except KeyError as e:
                    print("scheduler_id_format中存在异常%s" % e)
                else:
                    scheduler_id_list.append(scheduler_id)
            else:
                print("flow: %s scheduler创建失败" % data["flowid"])
                # return None
        print("------create_new_scheduler(self)执行结束, 返回scheduler_id_list------\n")
        print('scheduler_id_list', scheduler_id_list)
        return scheduler_id_list

    def get_execution_info(self):
        """
        根据schedulers id 查询出execution id, name,
        创建scheduler后查询execution有延迟，需要加等待时间
        """
        print("------开始执行get_execution_info(self)------\n")
        scheduler_id_list = self.create_new_scheduler()
        if scheduler_id_list:
            e_info_list = []
            count = 1
            for scheduler_id in scheduler_id_list:
                print('开始第%d 次查询，查询scheduler id 为 %s 的execution info' % (count, scheduler_id))
                # 等待30S后查询
                time.sleep(20)
                # 若没有查到execution id， 需要再次查询
                e_info = self.get_e_finial_status(scheduler_id)
                e_info_list.append(e_info)
                print('e_info_list:', e_info_list)
                count += 1
            # print('查询得到的e_info_list', e_info_list)
            print("------get_execution_info(self)执行结束------\n")
            return e_info_list
        else:
            print("返回的scheduler_id_list为空", scheduler_id_list)
            return None

    def get_e_finial_status(self, scheduler_id):
        """
        根据get_execution_info(self)返回的scheduler  id,
        查询该scheduler的execution 状态
        """
        print("------开始执行get_e_finial_status(self, scheduler_id)------\n")
        if scheduler_id:
            execution_sql = 'select id, status_type, flow_id , flow_scheduler_id from merce_flow_execution where flow_scheduler_id = "%s" ' % scheduler_id
            time.sleep(10)
            select_result = self.ms.ExecuQuery(execution_sql)
            print(execution_sql)
            print('查询execution 结果：', select_result)
            # print("根据scheduler id %s 查询execution，查询结果 %s: " % (scheduler_id, select_result))
            if select_result:
                e_info = {}
                # 从查询结果中取值
                try:
                    e_id = select_result[0]["id"]
                    print(e_id)
                    e_info["e_id"] = e_id
                    e_info["flow_id"] = select_result[0]["flow_id"]
                    e_info["flow_scheduler_id"] = select_result[0]["flow_scheduler_id"]
                    e_info["e_final_status"] = select_result[0]["status_type"]
                except IndexError as e:
                    print("取值时报错 %s" % e)
                    raise e
                # else:
                #     # 对返回数据格式化
                #     e_status_format = dict_res(e_status)
                #     e_final_status = e_status_format["type"]
                # e_info["e_final_status"] = e_final_status  #
                # 将 execution id , flow_id和status组装成字典的形式并返回
                print("------get_e_finial_status(self, scheduler_id)执行成功，返回execution id和status------\n")
                return e_info
            else:
                # print("根据scheduler id: %s ,没有查找到execution" % scheduler_id)
                return None
        else:
            return None

    def check_out_put(self):
        """
        获取execution的id和状态,
        最终返回execution执行成功后的dataset id
        """
        print("------开始执行check_out_put(self)------\n")
        e_info_list = self.get_execution_info()
        print("------check_out_put中得到的e_info:------\n", type(e_info_list), e_info_list)
        # 返回的len(e_info)和 len(flow_id_list)相等时，数据无缺失，进行后续的判断
        # if len(e_info_list) == len(flow_id_list):  # 表行-1
        if len(e_info_list) == len(self.file_flowid_count()):  # 与填入表的flowid数一样
            sink_dataset_list = []
            for i in range(len(e_info_list)):
                print('e_info_list', e_info_list)
                sink_dataset_dict = {}
                try:
                    e_id = e_info_list[i]["e_id"]
                    e_final_status = e_info_list[i]["e_final_status"]
                    e_scheduler_id = e_info_list[i]["flow_scheduler_id"]
                    e_flow_id = e_info_list[i]["flow_id"]
                except:
                    print('请确认该flow的e_info_list:')
                else:
                    sink_dataset_dict["flow_id"] = e_flow_id
                    sink_dataset_dict["execution_id"] = e_id
                    sink_dataset_dict["flow_scheduler_id"] = e_scheduler_id
                    if e_id:
                        print("------开始对%s 进行状态的判断------\n" % e_id)
                        while e_final_status in ("READY", "RUNNING"):
                            print("------进入while循环------\n")
                            # 状态为 ready 或者 RUNNING时，再次查询e_final_status
                            print("------查询前等待5S------\n")
                            time.sleep(5)
                            # 调用get_e_finial_status(e_scheduler_id)再次查询状态
                            e_info = self.get_e_finial_status(e_scheduler_id)
                            # 对e_final_status 重新赋值
                            e_final_status = e_info["e_final_status"]
                            print("------再次查询后的e_final_status: %s------\n" %e_final_status)
                            # time.sleep(50)
                        if e_final_status == "FAILED":
                            print("execution %s 执行失败" % e_id)
                            sink_dataset_dict["e_final_status"] = e_final_status
                            sink_dataset_dict["o_dataset"] = ""
                            sink_dataset_list.append(sink_dataset_dict)
                            # continue
                        elif e_final_status == "KILLED":
                            print("execution %s 被杀死" % e_id)
                            sink_dataset_dict["e_final_status"] = e_final_status
                            sink_dataset_dict["o_dataset"] = ""
                            sink_dataset_list.append(sink_dataset_dict)
                            # continue
                        elif e_final_status == "SUCCEEDED":
                            # 成功后查询flow_execution_output表中的dataset, 即sink对应的输出dataset，取出dataset id 并返回该ID，后续调用预览接口查看数据`
                            # print("查询data_json_sql:")
                            sink_dataset_dict["e_final_status"] = e_final_status
                            print(e_final_status, e_id)
                            data_json_sql = 'select b.dataset_json from merce_flow_execution as a  LEFT JOIN merce_flow_execution_output as b on a.id = b.execution_id where a.id ="%s"' % e_id
                            # print(data_json_sql)
                            data_json = self.ms.ExecuQuery(data_json_sql)
                            # print("打印data_json:", data_json)
                            for n in range(len(data_json)):
                                sink_dataset = data_json[n]["dataset_json"]  # 返回结果为元祖
                                print('-----sink_dataset-----', sink_dataset)
                                # print('sink_dataset:', sink_dataset)
                                if sink_dataset:
                                    sink_dataset_id = dict_res(sink_dataset)["id"]  # 取出json串中的dataset id
                                    sink_dataset_dict["o_dataset"] = sink_dataset_id
                                    d = json.loads(json.dumps(sink_dataset_dict))
                                    sink_dataset_list.append(d)
                                else:
                                    continue

                            print('第%d次的sink_dataset_list %s' % (i, sink_dataset_list))
                        else:
                            print("返回的execution 执行状态不正确")
                            return
                    else:
                        print("execution不存在")
                        return
            print('最后返回的sink_dataset_list\n', sink_dataset_list)
            print("------check_out_put(self)执行结束------\n")
            # 调用get_json()方法获取dataset_json
            return sink_dataset_list
        else:
            print("返回的scheduler_id_list值缺失")
            return

    def get_json(self):
        print("------开始执行get_json()------\n")
        sink_dataset = self.check_out_put()
        print('打印sink_dataset', sink_dataset)
        sink_dataset_json = []
        # ---使用openpyxl处理表格 12.26update---
        flow_table = load_workbook(abs_dir("flow_dataset_info.xlsx"))
        # info_sheet_names = flow_table.get_sheet_names()
        flow_sheet = flow_table.get_sheet_by_name('flow_info')
        sheet_rows = flow_sheet.max_row  # 获取行数

        # 通过dataset预览接口取得数据的预览json串 result.text
        for i in range(0, len(sink_dataset)):
            dataset_id = sink_dataset[i]["o_dataset"]
            # 第二步：判断dataset id是否存在，存在则取回预览结果并找到表中相等的dataset id，写入预览结果
            # print('dataset_id: ', dataset_id)
            for j in range(2, sheet_rows + 1):  # 按照行数进行循环
                # print('dataset_id:', dataset_id)
                # 通过dataset预览接口，获取dataset json串
                priview_url = "%s/api/datasets/%s/preview?rows=5000&tenant=2d7ad891-41c5-4fba-9ff2-03aef3c729e5" % (
                HOST_189, dataset_id)
                result = requests.get(url=priview_url, headers=get_headers())
                # print('预览接口返回的dataset json串', '\n', result.json())
                # 如果dataset id相等就写入实际结果，不相等就向下找
                # if dataset_id:
                    # for n in range(j, sheet_rows+1):
                if dataset_id == flow_sheet.cell(row=j, column=4).value:
                    flow_sheet.cell(row=j, column=8, value=result.text)  # dataset id 相等，实际结果写入表格
                # flow id 相等时，将execution id 和执行状态写入
                if sink_dataset[i]["flow_id"] == flow_sheet.cell(row=j, column=2).value:
                    # print(sink_dataset[i]["flow_id"])
                    flow_sheet.cell(row=j, column=5, value=sink_dataset[i]["execution_id"])
                    flow_sheet.cell(row=j, column=6, value=sink_dataset[i]["e_final_status"])

                else:
                    for t in range(j, 2, -1):
                        if sink_dataset[i]["flow_id"] == flow_sheet.cell(row=t-1, column=2).value:
                            flow_sheet.cell(row=j, column=5, value=sink_dataset[i]["execution_id"])
                            flow_sheet.cell(row=j, column=6, value=sink_dataset[i]["e_final_status"])
                            # flow_sheet.cell(row=j, column=8, value=result.text)  # 实际结果写入表格
                            break
            # else:
            #     print('请确认flow id = %s 的dataset id' % sink_dataset[i]["flow_id"])

        flow_table.save(abs_dir("flow_dataset_info.xlsx"))
        # copy_table.save(abs_dir("flow_dataset_info.xlsx"))
        # ---使用openpyxl处理表格 12.26update---
        table = load_workbook(abs_dir("flow_dataset_info.xlsx"))
        table_sheet = table.get_sheet_by_name('flow_info')
        c_rows = table_sheet.max_row

        # mode = overwrite:实际结果写入表后，对比预期结果和实际结果,并把失败详情存在 fail_detail
        print('-----开始对比结果----')
        for i in range(2, c_rows+1):
            table_sheet.cell(row=i, column=1, value=i-1)
            # 对sampe step涉及的flow单独进行结果判断
            if table_sheet.cell(row=i, column=2).value in ('09296a11-5abf-4af4-a58f-2f14e414db67', '981b6f96-5106-4b4e-8b11-d3757d17baaf',  # 189 sample
                                                            'ee3a57bf-494e-4e03-9755-fc6ad1d22a2a', '2d095376-2e72-4169-9d9b-68ade8f40955'):  # 84 sample
                if table_sheet.cell(row=i, column=6).value == 'SUCCEEDED' and table_sheet.cell(row=i, column=8):
                    new_result = []
                    expect_result = list(eval(table_sheet.cell(row=i, column=7).value))
                    actual_result = list(eval(table_sheet.cell(row=i, column=8).value))
                    for b_item in range(len(expect_result)):
                        for a_item in range(len(actual_result)):
                            if actual_result[a_item] == expect_result[b_item]:
                                new_result.append(actual_result[a_item])
                    if len(new_result) == len(actual_result):
                        # print(new_result)
                        # print(actual_result)
                        # print(expect_result)
                        table_sheet.cell(row=i, column=9, value="pass")
                        print('test_result:', table_sheet.cell(row=i, column=9).value)
                        table_sheet.cell(row=i, column=10, value="")
                    else:
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="execution: %s 预期结果实际结果不一致 " %
                                                                 (table_sheet.cell(row=i, column=5).value))
                elif table_sheet.cell(row=i, column=6).value == 'SUCCEEDED' and table_sheet.cell(row=i, column=8) == "":
                    table_sheet.cell(row=i, column=9, value="fail")
                    table_sheet.cell(row=i, column=10, value="execution: %s 预期结果实际结果不一致,实际结果为空 " %
                                                             (table_sheet.cell(row=i, column=5).value))
                elif table_sheet.cell(row=i, column=6).value == 'FAILED':
                    table_sheet.cell(row=i, column=9, value="fail")
                    table_sheet.cell(row=i, column=10, value="execution: %s 执行状态为 %s" % (
                        table_sheet.cell(row=i, column=5).value, table_sheet.cell(row=i, column=6).value))
                else:
                    print('请确认flow_id: %s的执行状态' % table_sheet.cell(row=i, column=2).value)

            else:
                # 判断mode为overwrite
                if table_sheet.cell(row=i, column=11).value == 'overwrite':  # 判断mode
                    # 实际结果存在并且执行结果为succeeded
                    if table_sheet.cell(row=i, column=8).value and table_sheet.cell(row=i, column=6).value == "SUCCEEDED":
                        # va7为预期结果，va8为实际结果，将二者先排序后对比是否相等
                        va7 = list(eval(table_sheet.cell(row=i, column=7).value))
                        va8 = list(eval(table_sheet.cell(row=i, column=8).value))   # 注意：是不是需要放在if 语句后面？？？？
                        if va7 != [] and eval(table_sheet.cell(row=i, column=8).value).__class__ == [].__class__ :
                            va7_k = va7[0].keys()
                            va7_key = list(va7_k)
                            # print('va7_key', va7_key)
                            S_va7 = sorted(va7, key=lambda item: item[va7_key[0]], reverse=True)    # 没有 id时候的排序
                            S_va8 = sorted(va8, key=lambda item: item[va7_key[0]], reverse=True)
                            # print('flow_id', table_sheet.cell(row=i, column=2).value)
                            # print(S_va7, '\n', S_va8)
                            # 安排不同的key进行排序，只要有其中一个key排序后相等，就认为两个结果相等
                            result = []
                            for t in range(len(va7_key)):
                                S_va7 = sorted(va7, key=lambda item: item[va7_key[t]], reverse=True)  # 没有 id时候的排序
                                S_va8 = sorted(va8, key=lambda item: item[va7_key[t]], reverse=True)
                                result.append(S_va7 == S_va8)
                            print('-----确认结果------')
                            if True in result:
                                table_sheet.cell(row=i, column=9, value="pass")
                                print('test_result:', table_sheet.cell(row=i, column=9).value)
                                table_sheet.cell(row=i, column=10, value="")
                            else:
                                table_sheet.cell(row=i, column=9, value="fail")
                                table_sheet.cell(row=i, column=10, value="flowname: %s --->预期结果实际结果不一致 \n" %
                                                                         (table_sheet.cell(row=i, column=3).value))
                        elif va7 == [] and va8 == []:
                            table_sheet.cell(row=i, column=9, value="pass")
                            print('test_result:', table_sheet.cell(row=i, column=9).value)
                            table_sheet.cell(row=i, column=10, value="")
                        else:
                            table_sheet.cell(row=i, column=9, value="")
                            table_sheet.cell(row=i, column=10, value="请确认预期结果和实际结果")
                    elif table_sheet.cell(row=i, column=6).value == "FAILED":
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="flowname: %s --->执行状态为 %s\n" % (table_sheet.cell(row=i, column=3).value, table_sheet.cell(row=i, column=6).value))
                    else:
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="用例参数或datasetID填写错误")
                # 判断mode为append
                elif table_sheet.cell(row=i, column=11).value == 'append':
                    if table_sheet.cell(row=i, column=8).value and table_sheet.cell(row=i, column=6).value == "SUCCEEDED":  # 实际结果存在
                        expect_result_list = list(eval(table_sheet.cell(row=i, column=7).value))
                        expect_len = len(expect_result_list)
                        actual_result_list = list(eval(table_sheet.cell(row=i, column=8).value))

                        if expect_result_list == actual_result_list[-expect_len:]:  # 实际结果切片和预期结果长度一致的数据，判断和预期结果是否相等
                            # print('expect_result_list:', expect_result_list)
                            # print('actual_result_list:', actual_result_list)
                            # print(expect_result_list == actual_result_list[-expect_len:])
                            table_sheet.cell(row=i, column=9, value="pass")
                            table_sheet.cell(row=i, column=10, value="")
                        else:
                            table_sheet.cell(row=i, column=9, value="fail")
                            table_sheet.cell(row=i, column=10,
                                                   value="execution: %s 预期结果实际结果不一致 \n预期结果: %s\n实际结果: %s" % (
                                                   table_sheet.cell(row=i, column=5).value,
                                                   table_sheet.cell(row=i, column=7).value,
                                                   table_sheet.cell(row=i, column=8).value))
                    elif table_sheet.cell(row=i, column=6).value == "FAILED":  # execution执行失败
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10,
                                               value="execution: %s 执行状态为 %s" % (
                                               table_sheet.cell(row=i, column=5).value, table_sheet.cell(row=i, column=6).value))
                    else:
                        table_sheet.cell(row=i, column=9, value="fail")
                        table_sheet.cell(row=i, column=10, value="用例参数或datasetID填写错误")

                else:
                    table_sheet.cell(row=i, column=9, value="fail")
                    table_sheet.cell(row=i, column=10, value="请确认flow的mode")
        # print(table_sheet.cell(row=2, column=8).value)
        # print(table_sheet.cell(row=2, column=9).value)
        table.save(abs_dir("flow_dataset_info.xlsx"))
        print('结果保存结束')


if __name__ == '__main__':
    # GetCheckoutDataSet()
    g = GetCheckoutDataSet()
    g.get_json()
    # g.get_json()
    # begin_time = datetime.datetime.now()
    # print('begin_time:', begin_time)
    # g = GetCheckoutDataSet()
    # print(g.get_execution_info())
    # end_time = datetime.datetime.now()
    # print('end_time:', end_time)
    # print('此次执行耗时：', end_time-begin_time)


    # threading.Timer(1500, get_headers()).start()
    # g.data_for_create_scheduler()



