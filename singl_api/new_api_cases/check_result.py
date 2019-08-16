# coding:utf-8
import re

from openpyxl import load_workbook
import os,unittest
from util.format_res import dict_res

# table_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))
# case_table = load_workbook(table_dir('api_cases_83.xlsx'))
# case_table_sheet = case_table.get_sheet_by_name('tester')
# all_rows = case_table_sheet.max_row

# 对比code和text
class CheckResult(unittest.TestCase):

    def __init__(self, table_name, case_table, case_table_sheet, all_rows):
        self.ab_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))
        self.table_name = table_name
        self.case_table = case_table
        self.case_table_sheet = case_table_sheet
        self.all_rows = all_rows
    def compare_code_result(self):
        """1.对比预期code和接口响应返回的status code"""
        for row in range(2, self.all_rows+1):
            # 预期status code和接口返回status code
            ex_status_code = self.case_table_sheet.cell(row=row, column=7).value
            ac_status_code = self.case_table_sheet.cell(row=row, column=8).value
            # 判断两个status code是否相等
            if ex_status_code and ac_status_code != '':
                # code相等时，pass
                if ex_status_code == ac_status_code:
                    self.case_table_sheet.cell(row=row, column=9, value='pass')
                else:
                    self.case_table_sheet.cell(row=row, column=9, value='fail') # code不等时，用例结果直接判断为失败
                    print('预期结果：%s, 实际结果：%s' % (ex_status_code, ac_status_code))
            else:
                print('第 %d 行 status_code为空' % row)
        self.case_table.save(self.ab_dir('api_cases_83.xlsx'))

    # 对比预期response和实际返回的response.text，根据预期和实际结果的关系进行处理
    def compare_text_result(self):
        for row in range(2, self.all_rows+1):
            response_text = self.case_table_sheet.cell(row=row, column=12).value  # 接口返回的response.text
            response_text_dict = dict_res(response_text)
            expect_text = self.case_table_sheet.cell(row=row, column=10).value  # 预期结果
            key_word = self.case_table_sheet.cell(row=row, column=3).value  # 接口关键字
            code_result = self.case_table_sheet.cell(row=row, column=9).value  # status_code对比结果
            relation = self.case_table_sheet.cell(row=row, column=11).value  # 预期text和response.text的关系
            #  1.status_code 对比结果pass的前提下，判断response.text断言是否正确,
            #  2.status_code 对比结果fail时，用例整体结果设为fail
            if code_result == 'pass':
                if key_word in ('create', 'query', 'update', 'delete'):
                    self.assert_deal(key_word, relation, expect_text, response_text, response_text_dict, row, 13)
                else:
                    print('请确认第%d行的key_word' % row)
            elif code_result == 'fail':
                # case 结果列
                self.case_table_sheet.cell(row=row, column=14, value='fail')
                # case失败原因
                self.case_table_sheet.cell(row=row, column=15, value='status_code对比结果为%s' % code_result)
            else:
                print('请确认第 %d 行 status_code对比结果' % row)

        self.case_table.save(self.ab_dir('api_cases_83.xlsx'))

    #  根据expect_text, response_text的关系，进行断言, 目前只处理了等于和包含两种关系
    def assert_deal(self, key_word, relation, expect_text, response_text, response_text_dict, row, column):
        if key_word == 'create':
            if relation == '=':   # 只返回id时，判断返回内容中包含id属性，id长度为36
                if isinstance(response_text_dict, dict):
                    if response_text_dict.get("id"):
                        # 返回的内容中包含 id属性，判断返回的id长度和预期给定的id长度一致
                        try:
                            self.assertEqual(expect_text, len(response_text_dict['id']), '第%d行的response_text长度和预期不一致' % row)
                        except:
                            print('第 %d 行 response_text返回的id和预期id长度不一致' % row)
                            self.case_table_sheet.cell(row=row, column=column, value='fail')
                        else:
                            self.case_table_sheet.cell(row=row, column=column, value='pass')
                    else:
                        try:
                            self.assertEqual(expect_text, response_text, '第%d行的response_text长度和预期不一致' % row)
                        except:
                            print('第 %d 行 response_text和预期text不相等' % row)
                            self.case_table_sheet.cell(row=row, column=column, value='fail')
                        else:
                            self.case_table_sheet.cell(row=row, column=column, value='pass')
                else:  # 只返回一个id串的情况下，判断预期长度和id长度一致
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%d行的response_text长度和预期不一致' % row)
                    except:
                        print('第 %d 行 response_text和预期text不相等' % row)
                        self.case_table_sheet.cell(row=row, column=column, value='fail')
                    else:
                        self.case_table_sheet.cell(row=row, column=column, value='pass')

            elif relation == 'in':  # 返回多内容时，判断返回内容中包含id属性，并且expect_text包含在response_text中
                try:
                    # self.assertIsNotNone(response_text_dict.get("id"), '第 %d 行 response_text没有返回id' % row)
                    self.assertIn(expect_text, response_text, '第 %d 行 expect_text没有包含在接口返回的response_text中' % row)
                except:
                    print('第 %d 行 expect_text没有包含在response_text中， 结果对比失败' % row)
                    self.case_table_sheet.cell(row=row, column=column, value='fail')
                else:
                    self.case_table_sheet.cell(row=row, column=column, value='pass')
            else:
                print('请确认第 %d 行 预期expect_text和response_text的relatrion' % row)
                self.case_table_sheet.cell(row=row, column=column, value='请确认预期text和接口response.text的relatrion')
        elif key_word in ('query', 'update', 'delete'):
            if relation == '=':
                compare_result = re.findall('[a-z0-9]{8}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{4}-[a-z0-9]{12}', '%s' % (response_text))
                response_text_list = []
                response_text_list.append(response_text)
                # 返回值是id 串，字母和数字的组合
                if compare_result == response_text_list:
                    try:
                        self.assertEqual(expect_text, len(response_text), '第%s行expect_text和response_text不相等' % row)
                    except:
                        self.case_table_sheet.cell(row=row, column=column, value='fail')
                    else:
                        self.case_table_sheet.cell(row=row, column=column, value='pass')
                # 返回空值
                elif expect_text == None and response_text == "":
                    self.case_table_sheet.cell(row=row, column=column, value='pass')

                else:
                    try:
                        self.assertEqual(expect_text, response_text, '第%s行expect_text和response_text不相等' % row)
                    except:
                        self.case_table_sheet.cell(row=row, column=column, value='fail')
                    else:
                        self.case_table_sheet.cell(row=row, column=column, value='pass')

            elif relation == 'in':
                try:
                    self.assertIn(expect_text, response_text, '第 %d 行 expect_text没有包含在接口返回的response_text中' % row)
                except:
                    print('第 %d 行 expect_text和response_text不相等， 结果对比失败' % row)
                    self.case_table_sheet.cell(row=row, column=column, value='fail')
                else:
                    self.case_table_sheet.cell(row=row, column=column, value='pass')
            else:
                print('请确认第 %d 行 预期expect_text和response_text的relatrion' % row)
                self.case_table_sheet.cell(row=row, column=column, value='请确认预期text和接口response.text的relatrion')
        else:
            print('请确认第 %d 行 的key_word' % row)
        self.case_table.save(self.ab_dir('api_cases_83.xlsx'))
    # 对比case最终的结果
    def deal_result(self):
        # 执行测试用例
        # deal_request_method()
        # 对比code
        self.compare_code_result()
        # 对比text
        self.compare_text_result()
        # 根据code result和text result判断case最终结果
        for row in range(2, self.all_rows + 1):
            status_code_result = self.case_table_sheet.cell(row=row, column=9).value
            response_text_result = self.case_table_sheet.cell(row=row, column=13).value
            if status_code_result == 'pass' and response_text_result == 'pass':
                # print('测试用例:%s 测试通过' % case_table_sheet.cell(row=row, column=3).value)
                self.case_table_sheet.cell(row=row, column=14, value='pass')
                self.case_table_sheet.cell(row=row, column=15, value='')
            elif status_code_result == 'fail' and response_text_result == 'pass':
                self.case_table_sheet.cell(row=row, column=14, value='fail')
                self.case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：status code对比失败,预期为%s,实际为%s'
                                                                % (self.case_table_sheet.cell(row=row, column=2).value, case_table_sheet.cell(row=row, column=7).value, case_table_sheet.cell(row=row, column=8).value))
            elif status_code_result == 'pass' and response_text_result == 'fail':
                self.case_table_sheet.cell(row=row, column=14, value='fail')
                self.case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：返回内容对比失败' %
                                                                (self.case_table_sheet.cell(row=row, column=2).value))
            elif status_code_result == 'fail' and response_text_result == 'fail':
                self.case_table_sheet.cell(row=row, column=14, value='fail')
                self.case_table_sheet.cell(row=row, column=15, value='%s--->失败原因：status code和返回文本对比均失败，请查看附件<api_cases_83.xlsx>确认具体失败原因'
                                                                % (self.case_table_sheet.cell(row=row, column=2).value))
            else:
                print('请确认status code或response.text对比结果')
        self.case_table.save(self.ab_dir('api_cases_83.xlsx'))


