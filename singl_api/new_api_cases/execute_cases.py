# coding:utf-8
import os
import re
import time
# from selenium import webdriver
from util import get_host
from openpyxl import load_workbook
import requests
from basic_info.get_auth_token import get_headers, get_headers_root,get_auth_token
from util.format_res import dict_res, get_time
from basic_info.setting import MySQL_CONFIG
from util.Open_DB import MYSQL
from new_api_cases.operate_request_method import deal_request_method


table_names = ["api_cases_83.xlsx", "api_cases_57.xlsx"]
ab_dir = lambda n: os.path.abspath(os.path.join(os.path.dirname(__file__), n))
jar_dir = ab_dir('woven-common-3.0.jar')


class ExecuteCases:
    def __init__(self):
        self.ms = MYSQL(MySQL_CONFIG["HOST"], MySQL_CONFIG["USER"], MySQL_CONFIG["PASSWORD"], MySQL_CONFIG["DB"])

    def execute(self):
        for table_name in table_names:
            case_table = load_workbook(ab_dir(table_name))
            case_table_sheet = case_table.get_sheet_by_name('tester')
            all_rows = case_table_sheet.max_row
            # 根据请求方法进行分发
            deal_request_method(table_name, case_table, case_table_sheet, all_rows, self.ms)

print(ExecuteCases().execute())

